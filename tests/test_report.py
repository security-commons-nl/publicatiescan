"""Tests voor de rapportage: de HTML is het triage-document, Excel is volledig.

Laag-bevindingen zijn op een volledige historie tienduizenden rijen verwachte ruis;
als rijen maken ze de HTML tientallen MB's en onbruikbaar. Ze worden samengevat
(telling per soort), nooit stil weggelaten. Excel houdt altijd alles.
"""
import os
import tempfile

from avgscan import report

_RIJEN = [
    ("https://x/besluit.pdf", "/tmp/a.pdf", "Persoonsnaam (aanhef)", "Hoog",
     "A. V*******d", "pagina 1", "De heer A. V*******d T*****aat 1", "naam bij woonadres"),
    ("https://x/form.pdf", "/tmp/d.pdf", "Telefoon (mobiel)", "Middel",
     "06*****99", "pagina 1", "mobiel 06*****99", "afwijkend nummer"),
    ("https://x/kennisgeving.pdf", "/tmp/b.pdf", "NAW (postcode+huisnr)", "Laag",
     "2*** AB", "pagina 1", "het perceel T*****aat 1, 2*** AB", "onderwerp-adres"),
    ("https://x/brief.pdf", "/tmp/c.pdf", "Telefoon (vast)", "Laag",
     "07*****00", "pagina 2", "bel 07*****00", "kan een zaaknummer zijn"),
]


def _html():
    d = tempfile.mkdtemp()
    pad = os.path.join(d, "rapport.html")
    report.write_html(_RIJEN, pad, scanned_files=3, scanned_pages=0)
    return open(pad, encoding="utf-8").read()


def test_html_toont_kritiek_hoog_als_rij_en_middel_laag_als_telling():
    doc = _html()
    assert "A. V*******d" in doc                      # Hoog-triage-rij staat erin
    assert "07*****00" not in doc                     # Laag-rij niet...
    assert "Telefoon (vast): 1" in doc                # ...maar de Laag-telling wel
    assert "06*****99" not in doc                     # Middel-rij niet als rij...
    assert "Telefoon (mobiel): 1" in doc              # ...maar de Middel-telling wel
    assert "rapport.xlsx" in doc                      # en de verwijzing naar volledig


def test_html_totaaltelling_blijft_alle_bevindingen():
    doc = _html()
    assert "<b>4</b> bevindingen" in doc              # 1 Hoog + 1 Middel + 2 Laag


def test_intro_komt_bovenaan_als_meegegeven():
    d = tempfile.mkdtemp()
    pad = os.path.join(d, "rapport.html")
    report.write_html(_RIJEN, pad, 3, 0,
                      intro_html="<h2>Bronnen</h2><p>KOOP SRU-API</p>")
    doc = open(pad, encoding="utf-8").read()
    assert '<div class="intro">' in doc
    assert "KOOP SRU-API" in doc
    # de intro staat vóór de bevindingentabel
    assert doc.index("KOOP SRU-API") < doc.index("<table>")


def test_geen_intro_geen_leeg_blok():
    d = tempfile.mkdtemp()
    pad = os.path.join(d, "rapport.html")
    report.write_html(_RIJEN, pad, 3, 0)
    assert '<div class="intro">' not in open(pad, encoding="utf-8").read()


def test_extra_kolommen_html_en_excel(tmp_path):
    import openpyxl
    # rijen met 3 extra velden achter de 8 basisvelden
    rijen = [r + ("Leiden", "dakkapel", "Omgevingsvergunning") for r in _RIJEN]
    hpad = str(tmp_path / "r.html")
    report.write_html(rijen, hpad, 4, 0, extra_headers=["Gemeente", "Onderwerp", "Vergunning"])
    doc = open(hpad, encoding="utf-8").read()
    assert "<th>Gemeente</th>" in doc and "<th>Vergunning</th>" in doc
    assert "Omgevingsvergunning" in doc            # extra kolom bij de Hoog-rij
    xpad = str(tmp_path / "r.xlsx")
    report.write_excel(rijen, xpad, extra_headers=["Gemeente", "Onderwerp", "Vergunning"])
    ws = openpyxl.load_workbook(xpad).active
    assert ws.cell(1, 9).value == "Gemeente" and ws.cell(1, 11).value == "Vergunning"


def test_excel_houdt_alles():
    import openpyxl
    d = tempfile.mkdtemp()
    pad = os.path.join(d, "rapport.xlsx")
    report.write_excel(_RIJEN, pad)
    ws = openpyxl.load_workbook(pad).active
    assert ws.max_row == 5                            # kop + 4 bevindingen


# ----------------------------------------------------------------------------------
# XML-ongeldige tekens uit de PDF-tekstlaag
#
# Aanleiding: een run van 2847 documenten schreef rapport.html goed weg en klapte
# daarna op wb.save() met "All strings must be XML compatible". _STUURTEKENS dekte de
# C0-stuurtekens wel, maar niet de surrogates (\ud800-\udfff) en noncharacters
# (\ufffe, \uffff) die uit PDF's met een kapotte tekstcodering komen. Eén zo'n teken
# in één bevinding kostte het hele Excel-bestand.
# ----------------------------------------------------------------------------------
import pytest

from avgscan.report import _excel_veilig


def _rij(context):
    """Minimale bevinding-rij met `context` als veld uit de tekstlaag."""
    return ("https://x/1", "/tmp/a.pdf", "BSN", "Kritiek", "12*****82",
            "pagina 3", context, "")


@pytest.mark.parametrize("teken, naam", [
    ("\x00", "NULL"),
    ("\x07", "bell"),
    ("\x0b", "vertical tab"),
    ("\x1f", "unit separator"),
    ("\ud800", "surrogate (laag)"),
    ("\udfff", "surrogate (hoog)"),
    ("\ufffe", "noncharacter fffe"),
    ("\uffff", "noncharacter ffff"),
])
def test_xml_ongeldige_tekens_worden_gestript(teken, naam):
    assert teken not in _excel_veilig(f"abc{teken}def")


@pytest.mark.parametrize("tekst", [
    "Raadsplein: Agenda en raadsstukken",
    "Kröller-Müller, café, ĳsselmeer",        # accenten en ligaturen
    "regel1\nregel2\ttab\rretour",             # tab, LF en CR zijn geldig in XML
    "emoji \U0001F600 buiten de BMP",
    "",
])
def test_geldige_tekst_blijft_ongemoeid(tekst):
    assert _excel_veilig(tekst) == tekst


def test_niet_strings_gaan_ongewijzigd_door():
    for waarde in (None, 42, 3.14, True):
        assert _excel_veilig(waarde) is waarde


def test_write_excel_met_kapotte_tekstlaag(tmp_path):
    """Regressietest: deze rijen lieten wb.save() eerder klappen."""
    import openpyxl
    rijen = [
        _rij("kapotte tekstlaag: \ud800 en \ufffe en \uffff"),
        _rij("NULL\x00 bell\x07 unit\x1f"),
        _rij("normaal: Kröller-Müller, café, 😀"),
    ]
    pad = str(tmp_path / "rapport.xlsx")
    report.write_excel(rijen, pad)                   # mag geen ValueError geven
    ws = openpyxl.load_workbook(pad).active
    assert ws.max_row == len(rijen) + 1
    teksten = [c.value or "" for c in ws["G"][1:]]
    assert any("kapotte tekstlaag" in t for t in teksten)   # leesbare inhoud bewaard
    assert any("Kröller-Müller" in t for t in teksten)
    assert not any("\ud800" in t or "\ufffe" in t or "\x00" in t for t in teksten)


def test_extra_kolommen_worden_ook_geschoond(tmp_path):
    import openpyxl
    rijen = [_rij("normaal") + ("Raadsplein \ufffe 2013", "Ingekomen brief", "")]
    pad = str(tmp_path / "rapport.xlsx")
    report.write_excel(rijen, pad, extra_headers=["Gemeente", "Onderwerp", "Vergunning"])
    ws = openpyxl.load_workbook(pad).active
    assert "\ufffe" not in (ws.cell(row=2, column=9).value or "")
