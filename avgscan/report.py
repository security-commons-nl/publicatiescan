"""Rapportage: HTML (printbaar, huisstijl) en Excel. Waarden zijn al gemaskeerd."""
from __future__ import annotations

import html
import os
import re
from datetime import date

from .detect import SEVERITY_ORDER

_ERNST_KLEUR = {"Kritiek": "#c8102e", "Hoog": "#c05600", "Middel": "#8a6d00", "Laag": "#5a5a5a"}

# Vergunning-/documenttype afgeleid uit de bekendmaking-titel, zodat een bevinding naar
# het juiste team te routeren is. Eerste match wint; volgorde = specifiek vóór generiek.
_VERGUNNING = [
    ("omgevingsvergunning", "Omgevingsvergunning"), ("kapvergunning", "Kapvergunning"),
    ("kappen", "Kapvergunning"), ("evenement", "Evenementenvergunning"),
    ("exploitatie", "Exploitatievergunning"), ("standplaats", "Standplaatsvergunning"),
    ("ligplaats", "Ligplaatsvergunning"), ("gehandicaptenparkeer", "Gehandicaptenparkeerplaats"),
    ("drank", "Drank/horeca"), ("horeca", "Drank/horeca"), ("uitweg", "Uitweg/inrit"),
    ("inrit", "Uitweg/inrit"), ("sloop", "Sloopmelding"), ("verkeersbesluit", "Verkeersbesluit"),
    ("verkeer", "Verkeer"), ("bestemmingsplan", "Bestemmingsplan"), ("subsidie", "Subsidie"),
    ("melding", "Melding"), ("vergunning", "Vergunning (overig)"),
]


def vergunning_type(titel: str) -> str:
    """Leid het vergunning-/documenttype af uit een titel; '' als niets herkend."""
    low = (titel or "").lower()
    for kw, label in _VERGUNNING:
        if kw in low:
            return label
    return ""


def _sorted(rows):
    # rows: (url, local_path, soort, ernst, waarde_masked, locatie, context, opmerking)
    return sorted(rows, key=lambda r: (SEVERITY_ORDER.get(r[3], 9), r[2], r[0]))


def write_html(rows, out_path, scanned_files, scanned_pages, intro_html=None, extra_headers=None):
    """Schrijf het HTML-rapport. `intro_html` (optioneel) is een blok ruwe HTML dat
    bovenaan komt: bedoeld voor een organisatie-specifieke management-samenvatting
    (bronnen, werkwijze, duiding). Het staat los van de tool zodat het rapport zelf
    generiek en herbruikbaar blijft; avg_scan vult het vanuit output_dir/intro.html.

    `extra_headers` (optioneel): kolomtitels voor extra velden. Rijen mogen dan langer
    zijn dan de 8 basisvelden; alles ná index 8 wordt als extra kolom getoond (bv.
    gemeente, onderwerp, vergunningtype voor routing naar het juiste team)."""
    extra_headers = extra_headers or []
    rows = _sorted(rows)
    tel = {}
    for r in rows:
        tel[r[3]] = tel.get(r[3], 0) + 1
    chips = " · ".join(f"{k}: {v}" for k, v in sorted(tel.items(), key=lambda kv: SEVERITY_ORDER.get(kv[0], 9)))

    # De HTML is het PRIORITEITSdocument: alleen Kritiek en Hoog als rijen — dat zijn de
    # vermoedelijke lekken (BSN, naam+woonadres) die stuk voor stuk beoordeeld moeten
    # worden. Middel en Laag bestaan op een volledige historie uit tienduizenden rijen
    # (afwijkende adressen, aanhef-namen, e-mail, verwachte ruis); als rijen maken die
    # het bestand tientallen MB's groot en onbruikbaar in een browser. Ze worden daarom
    # per ernst als telling-per-soort getoond, met de volledige lijst in het Excel-rapport
    # — samengevat, nooit stil weggelaten. Wie op adres-niveau wil triëren gebruikt Excel.
    SAMENVATTEN = ("Middel", "Laag")
    samengevat = {e: [r for r in rows if r[3] == e] for e in SAMENVATTEN}
    rows = [r for r in rows if r[3] not in SAMENVATTEN]

    def _per_soort(lst):
        t = {}
        for r in lst:
            t[r[2]] = t.get(r[2], 0) + 1
        return " · ".join(f"{k}: {v}" for k, v in sorted(t.items(), key=lambda kv: -kv[1]))

    samenvatting_html = "".join(
        f'<br><b>{e} ({len(samengevat[e])}):</b> '
        f'<span class="sub">{html.escape(_per_soort(samengevat[e])) or "geen"}</span>'
        for e in SAMENVATTEN if samengevat[e])
    n_samengevat = sum(len(v) for v in samengevat.values())

    trs = []
    for r in rows:
        url, path, soort, ernst, waarde, loc, ctx, opm = r[:8]
        extra = r[8:]
        kleur = _ERNST_KLEUR.get(ernst, "#5a5a5a")
        bestand = html.escape(os.path.basename(path or ""))
        extra_td = "".join(f"<td>{html.escape(str(x) if x is not None else '')}</td>" for x in extra)
        trs.append(
            f"<tr>"
            f"<td class='ernst' style='color:{kleur}'>{html.escape(ernst)}</td>"
            f"<td>{html.escape(soort)}</td>"
            f"<td class='mono'>{html.escape(waarde)}</td>"
            f"<td>{html.escape(loc)}</td>"
            f"<td><a href='{html.escape(url)}'>{html.escape(url)}</a><div class='sub'>{bestand}</div></td>"
            f"<td class='ctx'>{html.escape(ctx)}</td>"
            f"<td class='sub'>{html.escape(opm)}</td>"
            f"{extra_td}"
            f"</tr>"
        )
    extra_th = "".join(f"<th>{html.escape(h)}</th>" for h in extra_headers)

    doc = f"""<!doctype html><html lang="nl"><head><meta charset="utf-8">
<title>AVG Publicatie Scanner — bevindingen {date.today().isoformat()}</title>
<style>
 body{{font-family:-apple-system,"Segoe UI",Roboto,Arial,sans-serif;color:#1a1a1a;background:#f4f4f1;margin:0;line-height:1.4;font-size:13px;}}
 main{{max-width:1400px;margin:0 auto;padding:22px 26px 60px;}}
 h1{{font-size:19px;margin:0 0 2px;}}
 .sub{{color:#5a5a5a;font-size:11px;}}
 .band{{height:5px;background:#c8102e;border-radius:3px;margin:10px 0 14px;}}
 .meta{{background:#fff;border:1px solid #d6d6d6;border-left:4px solid #1f4e79;padding:9px 13px;border-radius:5px;margin-bottom:14px;font-size:12px;}}
 table{{border-collapse:collapse;width:100%;background:#fff;border:1px solid #d6d6d6;border-radius:6px;overflow:hidden;}}
 th,td{{text-align:left;padding:6px 9px;border-bottom:1px solid #ececec;vertical-align:top;font-size:11.5px;}}
 th{{background:#eef1f5;color:#1f4e79;text-transform:uppercase;font-size:10px;letter-spacing:.03em;position:sticky;top:0;}}
 .ernst{{font-weight:700;white-space:nowrap;}}
 .mono{{font-family:Consolas,monospace;}}
 .ctx{{color:#333;max-width:320px;}}
 td .sub{{color:#8a8a8a;font-size:10px;}}
 a{{color:#1f4e79;word-break:break-all;}}
 .intro{{background:#fff;border:1px solid #d6d6d6;border-left:4px solid #c8102e;padding:14px 20px;border-radius:5px;margin-bottom:16px;font-size:12.5px;line-height:1.5;}}
 .intro h2{{font-size:15px;color:#1f4e79;margin:16px 0 5px;}}
 .intro h2:first-child{{margin-top:0;}}
 .intro table{{margin:6px 0;}}
 .intro td,.intro th{{padding:4px 8px;}}
 @media print{{body{{background:#fff;}}th{{position:static;}}}}
</style></head><body><main>
<h1>publicatiescan — bevindingen</h1>
<div class="sub">Intern werkdocument · bevat mogelijk persoonsgegevens · gegenereerd {date.today().isoformat()}</div>
<div class="band"></div>
{f'<div class="intro">{intro_html}</div>' if intro_html else ''}
<div class="meta">
 <b>{len(rows) + n_samengevat}</b> bevindingen in <b>{scanned_files}</b> geanalyseerde documenten
 ({scanned_pages} pagina's gecrawld).<br>Verdeling: {html.escape(chips) or "geen"}.<br>
 <b>Let op:</b> waarden zijn gemaskeerd; elke hit vergt handmatige context-beoordeling
 (bestuurder/burger/medewerker/leverancier · bewust gepubliceerd vs. datalek).
 BSN-hits zijn gevalideerd met de elfproef, IBAN met mod-97.<br>
 Hieronder alleen <b>Kritiek</b> en <b>Hoog</b> als rijen (de vermoedelijke lekken). Middel en
 Laag staan samengevat; de volledige lijst met alle rijen staat in rapport.xlsx.{samenvatting_html}
</div>
<table>
<tr><th>Ernst</th><th>Soort</th><th>Waarde (gemaskeerd)</th><th>Locatie</th><th>Bron</th><th>Context</th><th>Opmerking</th>{extra_th}</tr>
{''.join(trs) if trs else f'<tr><td colspan={7 + len(extra_headers)}>Geen Kritiek- of Hoog-bevindingen.</td></tr>'}
</table>
</main></body></html>"""
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(doc)


# Alles wat XML 1.0 NIET toestaat. Bewust een negatieve klasse (houd alleen geldige
# tekens over) in plaats van een opsomming van verboden tekens: dat dekt behalve de
# C0-stuurtekens ook surrogates (\ud800-\udfff) en noncharacters (\ufffe, \uffff), die
# uit PDF's met een kapotte tekstcodering komen. Die laatste twee categorieën lieten de
# xlsx-export klappen aan het eind van een uur analyseren, terwijl de HTML al geschreven
# was — de bevindingen waren er dus wel, maar het Excel-bestand niet.
# Geldig per XML 1.0: #x9 | #xA | #xD | [#x20-#xD7FF] | [#xE000-#xFFFD] | [#x10000-#x10FFFF]
_STUURTEKENS = re.compile(
    r"[^\x09\x0a\x0d\x20-\ud7ff\ue000-\ufffd\U00010000-\U0010ffff]")


def _excel_veilig(v):
    """Strip stuurtekens uit de PDF-tekstlaag; openpyxl weigert ze (IllegalCharacterError)
    en laat anders de hele rapportage klappen aan het eind van een lange run."""
    if not isinstance(v, str):
        return v
    return _STUURTEKENS.sub("", v)


def write_excel(rows, out_path, extra_headers=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    extra_headers = extra_headers or []
    rows = _sorted(rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bevindingen"
    kop = ["Ernst", "Soort", "Waarde (gemaskeerd)", "Locatie", "URL", "Bestand", "Context", "Opmerking"]
    kop += list(extra_headers)
    ws.append(kop)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
    for r in rows:
        url, path, soort, ernst, waarde, loc, ctx, opm = r[:8]
        basis = (ernst, soort, waarde, loc, url, os.path.basename(path or ""), ctx, opm)
        ws.append([_excel_veilig(v) for v in (basis + tuple(r[8:]))])
    for col, breedte in zip("ABCDEFGH", (10, 20, 22, 18, 55, 30, 45, 30)):
        ws.column_dimensions[col].width = breedte
    ws.freeze_panes = "A2"
    wb.save(out_path)
